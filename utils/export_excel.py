from io import BytesIO
from typing import Any, Dict

from openpyxl import Workbook

SECTION_ORDER = ["fixed", "mobile", "waste", "process", "electricity", "steam"]


def build_workbook_bytes(company: Dict[str, Any], activity: Dict[str, Any], inputs_2025: Dict[str, str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "6.활동데이터 입력"

    ws["A1"] = "회사명"
    ws["B1"] = company.get("companyName", "")
    ws["A2"] = "사업자등록번호"
    ws["B2"] = company.get("bizNo", "")
    ws["A3"] = "수신메일"
    ws["B3"] = company.get("email", "")

    row_cursor = 5

    for section in SECTION_ORDER:
        rows = activity.get("sections", {}).get(section, [])
        if not rows:
            continue

        ws[f"A{row_cursor}"] = section
        row_cursor += 1

        for idx, row in enumerate(rows, start=1):
            meta = row.get("facility") or row.get("incinerator") or row.get("processType") or ""
            sub = row.get("fuel") or row.get("wasteType") or row.get("processSubType") or ""
            unit = row.get("unit") or ""

            ws[f"A{row_cursor}"] = idx
            ws[f"B{row_cursor}"] = meta
            ws[f"C{row_cursor}"] = sub
            ws[f"D{row_cursor}"] = unit

            for m in range(1, 13):
                ws.cell(row=row_cursor, column=4 + m, value=inputs_2025.get(f"{section}_{idx}_{m}", ""))

            row_cursor += 1

        row_cursor += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
